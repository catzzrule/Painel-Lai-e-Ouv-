"""
Gerador de Relatório Mensal Estatístico - Ouvidoria MESP
========================================================
Lê os dados consolidados da Ouvidoria e exporta um arquivo Excel (.xlsx)
completo com abas de KPIs, Distribuição por Natureza, SLA por Mês, Ranking de Áreas e Alertas.

Salva em: 'public/relatorios/relatorio_mensal_ouvidoria.xlsx'

Execução:
  python gerar_relatorio_mensal.py
"""

import os
import sys
import json
from datetime import datetime

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


def gerar_relatorio():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    public_dir = os.path.join(script_dir, "public")
    relatorios_dir = os.path.join(public_dir, "relatorios")
    os.makedirs(relatorios_dir, exist_ok=True)

    json_path = os.path.join(public_dir, "dados_ouvidoria.json")
    if not os.path.exists(json_path):
        # Gerar o json primeiro caso não exista
        import subprocess
        subprocess.run([sys.executable, os.path.join(script_dir, "converter_auto_ouvidoria.py")])

    if not os.path.exists(json_path):
        print(f"[ERRO] Arquivo {json_path} não encontrado.")
        sys.exit(1)

    with open(json_path, "r", encoding="utf-8") as f:
        dados = json.load(f)

    caminho_excel = os.path.join(relatorios_dir, "relatorio_mensal_ouvidoria.xlsx")
    print(f"[*] Gerando relatório Excel em: {caminho_excel}...")

    # Criação do arquivo Excel formatado via openpyxl
    wb = openpyxl.Workbook()
    
    # 1. Aba Resumo Executivo
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Executivo"
    ws_resumo.views.sheetView[0].showGridLines = True

    # Estilos
    cor_azul = "1E3A8A"
    cor_azul_claro = "DBEAFE"
    cor_verde = "065F46"
    cor_verde_claro = "D1FAE5"
    cor_cinza = "F3F4F6"
    
    font_titulo = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
    font_subtitulo = Font(name="Calibri", size=11, italic=True, color="4B5563")
    font_secao = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_dado = Font(name="Calibri", size=11)
    font_destaque = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
    
    fill_secao = PatternFill(start_color=cor_azul, end_color=cor_azul, fill_type="solid")
    fill_header = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    fill_zebrado = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_kpi = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    borda_fina = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Cabeçalho do Relatório
    ws_resumo["B2"] = "MINISTÉRIO DO ESPORTE — OUVIDORIA GERAL"
    ws_resumo["B2"].font = font_titulo
    ws_resumo["B3"] = f"Relatório Gerencial e Estatístico Consolidado • Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    ws_resumo["B3"].font = font_subtitulo

    # Seção de Indicadores Chave (KPIs)
    ws_resumo.merge_cells("B5:E5")
    ws_resumo["B5"] = "INDICADORES CHAVE DE DESEMPENHO (KPIs)"
    ws_resumo["B5"].font = font_secao
    ws_resumo["B5"].fill = fill_secao
    ws_resumo["B5"].alignment = Alignment(horizontal="center", vertical="center")

    kpis = dados.get("kpis", {})
    kpis_tabela = [
        ("Total de Manifestações Registradas", kpis.get("total", 0), "manifestações"),
        ("Manifestações Concluídas (Resolvidas)", kpis.get("concluidas", 0), f"{kpis.get('percentual_concluidas', 0)}% de resolução"),
        ("Manifestações em Aberto (Tratamento)", kpis.get("em_aberto", 0), "aguardando resposta"),
        ("Cumprimento do Prazo de SLA", f"{kpis.get('taxa_no_prazo', 0)}%", "taxa de conformidade"),
        ("Tempo Médio de Resposta ao Cidadão", f"{kpis.get('media_dias_resposta', 0)} dias", "meta oficial: 20 dias"),
        ("Média de Folga em Relação ao Prazo Legal", f"{kpis.get('media_dias_folga', 0)} dias", "dias de antecedência"),
        ("Total de Elogios Registrados", kpis.get("total_elogios", 0), "reconhecimentos formais"),
        ("Total de Reclamações Registradas", kpis.get("total_reclamacoes", 0), "pontos de melhoria"),
        ("Total de Denúncias Recebidas", kpis.get("total_denuncias", 0), "fiscalização e apuração"),
    ]

    for idx, (rotulo, valor, detalhe) in enumerate(kpis_tabela, start=6):
        ws_resumo.cell(row=idx, column=2, value=rotulo).font = font_dado
        c_val = ws_resumo.cell(row=idx, column=3, value=valor)
        c_val.font = font_destaque
        c_val.alignment = Alignment(horizontal="center")
        ws_resumo.cell(row=idx, column=4, value=detalhe).font = font_subtitulo
        for c in range(2, 5):
            ws_resumo.cell(row=idx, column=c).border = borda_fina
            if idx % 2 == 0:
                ws_resumo.cell(row=idx, column=c).fill = fill_zebrado

    # 2. Aba Natureza das Manifestações
    ws_nat = wb.create_sheet(title="Natureza e Tipologia")
    ws_nat.views.sheetView[0].showGridLines = True
    ws_nat["B2"] = "DISTRIBUIÇÃO POR NATUREZA DA MANIFESTAÇÃO"
    ws_nat["B2"].font = font_titulo
    
    ws_nat.cell(row=4, column=2, value="Natureza / Tipo").font = font_header
    ws_nat.cell(row=4, column=2).fill = fill_header
    ws_nat.cell(row=4, column=3, value="Quantidade").font = font_header
    ws_nat.cell(row=4, column=3).fill = fill_header
    ws_nat.cell(row=4, column=4, value="% do Total").font = font_header
    ws_nat.cell(row=4, column=4).fill = fill_header

    total = kpis.get("total", 1)
    naturezas = dados.get("naturezas", {})
    for i, (nat, qtd) in enumerate(naturezas.items(), start=5):
        ws_nat.cell(row=i, column=2, value=nat).font = font_dado
        c_q = ws_nat.cell(row=i, column=3, value=qtd)
        c_q.font = font_destaque
        c_q.alignment = Alignment(horizontal="right")
        pct = round((qtd / total) * 100, 1)
        c_p = ws_nat.cell(row=i, column=4, value=f"{pct}%")
        c_p.font = font_dado
        c_p.alignment = Alignment(horizontal="right")
        for col in range(2, 5):
            ws_nat.cell(row=i, column=col).border = borda_fina

    # 3. Aba Evolução Mensal & SLA
    ws_mensal = wb.create_sheet(title="Evolução Mensal & SLA")
    ws_mensal.views.sheetView[0].showGridLines = True
    ws_mensal["B2"] = "HISTÓRICO MENSAL DE DEMANDAS E SLA"
    ws_mensal["B2"].font = font_titulo

    headers_m = ["Mês / Ano", "Total Demandas", "Tempo Médio (dias)", "Meta SLA (dias)"]
    for col_i, h in enumerate(headers_m, start=2):
        cell = ws_mensal.cell(row=4, column=col_i, value=h)
        cell.font = font_header
        cell.fill = fill_header

    mensal = dados.get("mensal", {})
    meses = mensal.get("meses", [])
    quantidades = mensal.get("quantidades", [])
    tempos = mensal.get("media_dias_resposta", [])

    for idx, m in enumerate(meses, start=5):
        q = quantidades[idx - 5] if idx - 5 < len(quantidades) else 0
        t = tempos[idx - 5] if idx - 5 < len(tempos) else 0
        ws_mensal.cell(row=idx, column=2, value=m).font = font_dado
        ws_mensal.cell(row=idx, column=3, value=q).font = font_destaque
        ws_mensal.cell(row=idx, column=4, value=t).font = font_dado
        ws_mensal.cell(row=idx, column=5, value=20).font = font_subtitulo
        for col in range(2, 6):
            ws_mensal.cell(row=idx, column=col).border = borda_fina

    # 4. Aba Alertas e Pontos de Atenção
    ws_alertas = wb.create_sheet(title="Alertas & Monitoramento")
    ws_alertas.views.sheetView[0].showGridLines = True
    ws_alertas["B2"] = "PONTOS DE ATENÇÃO E ALERTAS ATIVOS"
    ws_alertas["B2"].font = font_titulo

    ws_alertas.cell(row=4, column=2, value="Severidade").font = font_header
    ws_alertas.cell(row=4, column=2).fill = fill_header
    ws_alertas.cell(row=4, column=3, value="Título do Alerta").font = font_header
    ws_alertas.cell(row=4, column=3).fill = fill_header
    ws_alertas.cell(row=4, column=4, value="Descrição e Detalhes").font = font_header
    ws_alertas.cell(row=4, column=4).fill = fill_header

    alertas = dados.get("alertas", [])
    for idx, alt in enumerate(alertas, start=5):
        c_sev = ws_alertas.cell(row=idx, column=2, value=alt.get("tipo", "").upper())
        c_sev.font = Font(name="Calibri", size=11, bold=True)
        if alt.get("tipo") == "danger":
            c_sev.font = Font(name="Calibri", size=11, bold=True, color="991B1B")
        elif alt.get("tipo") == "warning":
            c_sev.font = Font(name="Calibri", size=11, bold=True, color="92400E")
        else:
            c_sev.font = Font(name="Calibri", size=11, bold=True, color="1E40AF")
            
        ws_alertas.cell(row=idx, column=3, value=alt.get("titulo", "")).font = font_destaque
        desc = alt.get("descricao", "") + " " + alt.get("detalhes", "")
        ws_alertas.cell(row=idx, column=4, value=desc).font = font_dado
        for col in range(2, 5):
            ws_alertas.cell(row=idx, column=col).border = borda_fina

    # Ajustar largura automática das colunas em todas as abas
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(caminho_excel)
    print(f"[OK] Relatório gerado com sucesso: {caminho_excel}")
    print(f"     Tamanho: {os.path.getsize(caminho_excel) / 1024:.1f} KB")


if __name__ == "__main__":
    gerar_relatorio()
