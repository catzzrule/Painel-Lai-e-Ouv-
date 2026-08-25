"""
Conversor Automático: SharePoint → JSON para o Painel da Ouvidoria
===================================================================
Baixa a planilha do SharePoint Online e gera 'dados.json'.

Uso:
  python converter_auto.py                    (usa link configurado abaixo)
  python converter_auto.py <URL_SHAREPOINT>   (usa link passado por argumento)

Configuração:
  Cole o link de compartilhamento do SharePoint na variável SHAREPOINT_LINK abaixo.
"""

import json
import os
import sys
import tempfile
import re
from collections import Counter, defaultdict
from datetime import datetime
from urllib.parse import urlparse, parse_qs

try:
    import requests
except ImportError:
    print("Erro: módulo 'requests' não encontrado.")
    print("Instale com: pip install requests")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Erro: módulo 'openpyxl' não encontrado.")
    print("Instale com: pip install openpyxl")
    sys.exit(1)


# ============================================================
# ⬇️  COLE SEU LINK DE COMPARTILHAMENTO DO SHAREPOINT AQUI  ⬇️
# ============================================================
SHAREPOINT_LINK = "https://mdsgov-my.sharepoint.com/:x:/g/personal/rayssa_vitoria_esporte_gov_br/IQBvZ6Cbt41oRJBD1QuPbvLBAfoOSMBzDpwUbAT797m4ggA?e=FaCchc"
# ============================================================


def converter_link_sharepoint(link):
    """
    Converte um link de compartilhamento do SharePoint para um link 
    de download direto do arquivo .xlsx.
    
    Suporta formatos:
    - https://dominio.sharepoint.com/:x:/s/Site/EaBcDeFg?e=XyZ
    - https://dominio.sharepoint.com/sites/Site/_layouts/15/download.aspx?...
    - https://dominio-my.sharepoint.com/personal/user/Documents/arquivo.xlsx
    """
    if "download.aspx" in link:
        return link  # Já é link de download
    
    # Formato padrão de compartilhamento: /:x:/ (Excel)
    # Transforma em download adicionando &download=1
    if "/:x:/" in link or "/:w:/" in link or "/:p:/" in link:
        # Remove parâmetros existentes e adiciona download=1
        if "?" in link:
            return link + "&download=1"
        else:
            return link + "?download=1"
    
    # Tenta adicionar download=1 como fallback
    if "sharepoint.com" in link:
        if "?" in link:
            return link + "&download=1"
        else:
            return link + "?download=1"
    
    return link


def baixar_planilha(link):
    """Baixa a planilha do SharePoint e retorna o caminho do arquivo temporário."""
    print("[*] Baixando planilha do SharePoint...")
    
    download_url = converter_link_sharepoint(link)
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }
    
    try:
        response = requests.get(download_url, headers=headers, allow_redirects=True, timeout=60)
        response.raise_for_status()
        
        # Verificar se recebeu um arquivo Excel válido
        content_type = response.headers.get("Content-Type", "")
        if "text/html" in content_type and len(response.content) < 5000:
            print("[!] O link retornou HTML em vez de um arquivo Excel.")
            print("    Isso pode significar que:")
            print("    1. O link de compartilhamento requer autenticação")
            print("    2. O link expirou")
            print("    3. O formato do link não é suportado")
            print()
            print("    Tente gerar um novo link com permissão 'Qualquer pessoa com o link'")
            sys.exit(1)
        
        # Salvar localmente para debug
        local_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Relatorio_baixado.xlsx")
        with open(local_path, "wb") as f:
            f.write(response.content)
        
        tamanho_kb = len(response.content) / 1024
        print(f"[OK] Planilha baixada com sucesso ({tamanho_kb:.1f} KB)")
        return local_path
        
    except requests.exceptions.RequestException as e:
        print(f"[ERRO] Falha ao baixar planilha: {e}")
        sys.exit(1)


def parse_date(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%Y %H:%M:%S", "%d/%m/%y", "%d/%m/%y %H:%M:%S", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


def carregar_planilha(caminho):
    """Carrega a planilha, detecta a linha de cabeçalhos e calcula campos faltantes se necessário."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    print(f"[*] Abas encontradas: {wb.sheetnames}")
    ws = wb.active
    print(f"[*] Aba ativa: {ws.title}")

    # 1. Detectar a linha de cabeçalho
    header_row_idx = 1
    headers = []
    for r in range(1, 11):  # Busca nas primeiras 10 linhas
        row_vals = [cell.value for cell in ws[r]]
        # Se contiver os principais cabeçalhos do Fala.BR
        has_sit = any(isinstance(v, str) and ("Situa" in v) for v in row_vals)
        has_nup = any(isinstance(v, str) and ("Nup" in v) for v in row_vals)
        if has_sit and has_nup:
            header_row_idx = r
            headers = [str(v).strip() if v is not None else f"Coluna_{i}" for i, v in enumerate(row_vals)]
            break
            
    if not headers:
        # Fallback para a primeira linha
        header_row_idx = 1
        headers = [str(cell.value).strip() if cell.value is not None else f"Coluna_{i}" for i, cell in enumerate(ws[1])]
        
    print(f"[*] Cabeçalhos identificados na linha {header_row_idx}: {headers[:15]}...")
    
    # Mapear chaves para facilitar busca
    def encontrar_coluna(termos):
        for h in headers:
            if any(t in h for t in termos):
                return h
        return None

    col_cadastro = encontrar_coluna(["Cadastro", "Data de Cadastro"])
    col_prazo = encontrar_coluna(["Prazo de Atendimento", "Prazo Recurso", "Prazo"])
    col_resposta = encontrar_coluna(["Data de Resposta", "Data Resposta"])
    col_situacao = encontrar_coluna(["Situação", "Situa"])
    col_nup = encontrar_coluna(["Nup", "NUP", "Protocolo"])

    # Verificar se as colunas calculadas existem
    tem_ano_mes = encontrar_coluna(["Ano-Mês Cadastro", "Ano-M"]) is not None
    tem_dias_resp = encontrar_coluna(["Dias para Resposta"]) is not None
    tem_dias_sla = encontrar_coluna(["Dias de Prazo (SLA)"]) is not None
    tem_dias_folga = encontrar_coluna(["Dias de Folga"]) is not None
    tem_status_prazo = encontrar_coluna(["Status do Prazo"]) is not None

    print(f"[*] Colunas calculadas presentes na planilha? "
          f"Ano-Mês: {tem_ano_mes}, Dias Resposta: {tem_dias_resp}, "
          f"SLA: {tem_dias_sla}, Folga: {tem_dias_folga}, Status: {tem_status_prazo}")

    registros = []

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, len(headers) + 1)]
        
        # Ignorar linhas totalmente vazias
        if all(v is None for v in row_vals):
            continue
            
        registro = dict(zip(headers, row_vals))
        
        # Ignorar registros vazios ou sem dados essenciais (como NUP ou Situação)
        nup_val = registro.get(col_nup) if col_nup else None
        sit_val = registro.get(col_situacao) if col_situacao else None
        if nup_val is None or str(nup_val).strip() == "" or sit_val is None or str(sit_val).strip() == "":
            continue
        
        # Obter datas parsed
        val_cadastro = parse_date(registro.get(col_cadastro)) if col_cadastro else None
        val_prazo = parse_date(registro.get(col_prazo)) if col_prazo else None
        val_resposta = parse_date(registro.get(col_resposta)) if col_resposta else None
        
        # Formatar as datas existentes para o formato YYYY-MM-DD no dicionário
        if col_cadastro and val_cadastro:
            registro[col_cadastro] = val_cadastro.strftime("%Y-%m-%d")
        if col_prazo and val_prazo:
            registro[col_prazo] = val_prazo.strftime("%Y-%m-%d")
        if col_resposta and val_resposta:
            registro[col_resposta] = val_resposta.strftime("%Y-%m-%d")
            
        # Calcular campos se não existirem
        if not tem_ano_mes and val_cadastro:
            registro["Ano-Mês Cadastro"] = val_cadastro.strftime("%Y-%m")
            
        if not tem_dias_resp:
            if val_resposta and val_cadastro:
                registro["Dias para Resposta"] = (val_resposta - val_cadastro).days
            else:
                registro["Dias para Resposta"] = None
                
        if not tem_dias_sla:
            if val_prazo and val_cadastro:
                registro["Dias de Prazo (SLA)"] = (val_prazo - val_cadastro).days
            else:
                registro["Dias de Prazo (SLA)"] = None
                
        if not tem_dias_folga:
            if val_prazo and val_resposta:
                registro["Dias de Folga/Atraso"] = (val_prazo - val_resposta).days
            else:
                registro["Dias de Folga/Atraso"] = None
                
        if not tem_status_prazo:
            sit_val = str(registro.get(col_situacao) or "").strip()
            # Se já está respondida
            if val_resposta:
                if val_prazo and val_resposta <= val_prazo:
                    registro["Status do Prazo"] = "Respondida no prazo"
                else:
                    registro["Status do Prazo"] = "Respondida fora do prazo"
            else:
                # Em aberto
                if val_prazo and datetime.now() > val_prazo:
                    registro["Status do Prazo"] = "Em atraso"
                else:
                    registro["Status do Prazo"] = "Em aberto - no prazo"

        registros.append(registro)

    if registros:
        print(f"[*] Exemplo de 1º registro processado: {list(registros[0].items())[:10]}")

    # Atualizar headers com os novos campos calculados
    headers_finais = list(headers)
    novos_campos = ["Ano-Mês Cadastro", "Dias para Resposta", "Dias de Prazo (SLA)", "Dias de Folga/Atraso", "Status do Prazo"]
    for campo in novos_campos:
        if campo not in headers_finais:
            headers_finais.append(campo)

    return registros, headers_finais


def _get_key(registro, substring):
    """Busca uma chave no registro que contenha a substring (para lidar com encoding)."""
    for k, v in registro.items():
        if k and substring in str(k):
            return v
    return None


def processar_dados(registros):
    """Processa os registros e gera os dados agregados para o painel."""
    dados = {}

    # --- Metadados ---
    dados["meta"] = {
        "total_registros": len(registros),
        "data_geracao": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    # --- KPIs ---
    situacoes = Counter()
    decisoes = Counter()
    subtipos = Counter()
    status_prazo = Counter()
    recursos = Counter()
    generos = Counter()
    tipos_pessoa = Counter()
    estados = Counter()
    meses = Counter()
    
    # Para capturar as áreas mais demandadas
    areas_situacoes = defaultdict(Counter)
    areas_decisoes = defaultdict(Counter)
    areas_subtipos = defaultdict(Counter)
    areas_status = defaultdict(Counter)
    areas_recursos = defaultdict(Counter)

    dias_resposta_total = []
    dias_prazo_total = []
    dias_folga_total = []

    # Agregações mensais
    mensal_qtd = defaultdict(int)
    mensal_dias_resp = defaultdict(list)
    mensal_situacao = defaultdict(lambda: Counter())
    mensal_decisao = defaultdict(lambda: Counter())
    mensal_status_prazo = defaultdict(lambda: Counter())

    for r in registros:
        sit = r.get("Situação") or r.get("Situa\u00e7\u00e3o") or _get_key(r, "Situa")
        subtipo = r.get("Subtipo de Formulário") or r.get("Subtipo de Formul\u00e1rio") or _get_key(r, "Subtipo")
        decisao = r.get("Especificação Decisão") or r.get("Especifica\u00e7\u00e3o Decis\u00e3o") or _get_key(r, "Especifica")
        status = r.get("Status do Prazo") or _get_key(r, "Status do Prazo")
        recurso = r.get("Situação do Recurso") or r.get("Situa\u00e7\u00e3o do Recurso") or _get_key(r, "Recurso")
        genero = r.get("Gênero") or r.get("G\u00eanero") or _get_key(r, "nero")
        tipo_p = r.get("Tipo de Pessoa") or _get_key(r, "Tipo de Pessoa")
        estado = r.get("Estado") or _get_key(r, "Estado")
        ano_mes = r.get("Ano-Mês Cadastro") or r.get("Ano-M\u00eas Cadastro") or _get_key(r, "Ano-M")
        dias_resp = r.get("Dias para Resposta") or _get_key(r, "Dias para Resposta")
        dias_prazo = r.get("Dias de Prazo (SLA)") or _get_key(r, "Dias de Prazo")
        dias_folga = r.get("Dias de Folga/Atraso") or _get_key(r, "Dias de Folga")
        
        # Área Responsável
        resp = str(r.get("Responsável pela Resposta") or _get_key(r, "Respons") or "Não Identificada").strip()

        if sit:
            situacoes[str(sit)] += 1
            if resp and resp != "None": areas_situacoes[str(sit)][resp] += 1
        if subtipo:
            subtipos[str(subtipo)] += 1
            if resp and resp != "None": areas_subtipos[str(subtipo)][resp] += 1
        if decisao:
            decisoes[str(decisao)] += 1
            if resp and resp != "None": areas_decisoes[str(decisao)][resp] += 1
        if status:
            status_prazo[str(status)] += 1
            if resp and resp != "None": areas_status[str(status)][resp] += 1
        if recurso:
            recursos[str(recurso)] += 1
            if resp and resp != "None": areas_recursos[str(recurso)][resp] += 1
        if genero:
            generos[str(genero)] += 1
        if tipo_p:
            tipos_pessoa[str(tipo_p)] += 1
        if estado:
            estados[str(estado)] += 1
        if ano_mes:
            mes_str = str(ano_mes)
            meses[mes_str] += 1
            mensal_qtd[mes_str] += 1
            if sit:
                mensal_situacao[mes_str][str(sit)] += 1
            if decisao:
                mensal_decisao[mes_str][str(decisao)] += 1
            if status:
                mensal_status_prazo[mes_str][str(status)] += 1

        if dias_resp is not None:
            try:
                val = int(dias_resp)
                dias_resposta_total.append(val)
                if ano_mes:
                    mensal_dias_resp[str(ano_mes)].append(val)
            except (ValueError, TypeError):
                pass
        if dias_prazo is not None:
            try:
                dias_prazo_total.append(int(dias_prazo))
            except (ValueError, TypeError):
                pass
        if dias_folga is not None:
            try:
                dias_folga_total.append(int(dias_folga))
            except (ValueError, TypeError):
                pass

    # KPIs
    total = len(registros)

    concluidas = 0
    for k, v in situacoes.items():
        if "onclu" in k:
            concluidas += v

    em_aberto = 0
    for k in situacoes:
        if "Cadastrada" in k or "Prorrogada" in k or "Encaminhada" in k:
            em_aberto += situacoes[k]

    com_recurso = 0
    for k in recursos:
        if "respondido" in k.lower():
            com_recurso += recursos[k]

    respondidas_prazo = 0
    for k in status_prazo:
        if "respondida" in k.lower() and "prazo" in k.lower():
            respondidas_prazo += status_prazo[k]

    no_prazo_total = 0
    for k in status_prazo:
        if "prazo" in k.lower():
            no_prazo_total += status_prazo[k]

    media_resp = round(sum(dias_resposta_total) / len(dias_resposta_total), 1) if dias_resposta_total else 0
    media_prazo = round(sum(dias_prazo_total) / len(dias_prazo_total), 1) if dias_prazo_total else 0
    media_folga = round(sum(dias_folga_total) / len(dias_folga_total), 1) if dias_folga_total else 0

    dados["kpis"] = {
        "total": total,
        "concluidas": concluidas,
        "percentual_concluidas": round((concluidas / total) * 100, 1) if total > 0 else 0,
        "em_aberto": em_aberto,
        "com_recurso": com_recurso,
        "respondidas_prazo": respondidas_prazo,
        "taxa_no_prazo": round((respondidas_prazo / concluidas) * 100, 1) if concluidas > 0 else 0,
        "no_prazo_total": no_prazo_total,
        "media_dias_resposta": media_resp,
        "media_dias_prazo_sla": media_prazo,
        "media_dias_folga": media_folga,
    }

    # --- Distribuições ---
    dados["situacoes"] = dict(situacoes.most_common())
    dados["decisoes"] = dict(decisoes.most_common())
    dados["subtipos"] = dict(subtipos.most_common())
    dados["status_prazo"] = dict(status_prazo.most_common())
    dados["recursos"] = dict(recursos.most_common())
    dados["generos"] = dict(generos.most_common())
    dados["tipos_pessoa"] = dict(tipos_pessoa.most_common())
    dados["estados"] = dict(sorted(estados.items(), key=lambda x: -x[1]))

    def extrair_top_area(counter_dict):
        resultado = {}
        for k, counter in counter_dict.items():
            top = counter.most_common(1)
            if top:
                resultado[k] = f"{top[0][0]} ({top[0][1]})"
            else:
                resultado[k] = "Não informada"
        return resultado

    dados["principais_areas"] = {
        "situacoes": extrair_top_area(areas_situacoes),
        "decisoes": extrair_top_area(areas_decisoes),
        "subtipos": extrair_top_area(areas_subtipos),
        "status_prazo": extrair_top_area(areas_status),
        "recursos": extrair_top_area(areas_recursos)
    }

    # --- Dados Mensais ---
    meses_ordenados = sorted(mensal_qtd.keys())
    dados["mensal"] = {
        "meses": meses_ordenados,
        "quantidades": [mensal_qtd[m] for m in meses_ordenados],
        "media_dias_resposta": [
            round(sum(mensal_dias_resp[m]) / len(mensal_dias_resp[m]), 1)
            if mensal_dias_resp[m]
            else 0
            for m in meses_ordenados
        ],
        "situacoes_por_mes": {m: dict(mensal_situacao[m]) for m in meses_ordenados},
        "status_prazo_por_mes": {m: dict(mensal_status_prazo[m]) for m in meses_ordenados},
    }

    # --- Lista detalhada (sem dados sensíveis) ---
    dados["registros"] = []
    for r in registros:
        registro_limpo = {}
        for chave, valor in r.items():
            # Omitir colunas sensíveis
            chave_lower = str(chave).lower() if chave else ""
            if any(
                termo in chave_lower
                for termo in ["usuário", "usuario", "responsável", "responsavel"]
            ):
                continue
            registro_limpo[chave] = valor
        dados["registros"].append(registro_limpo)

    return dados


def main():
    # Determinar a fonte dos dados
    script_dir = os.path.dirname(os.path.abspath(__file__))
    arquivo_temp = None
    
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        if arg.startswith("http"):
            # Argumento é uma URL do SharePoint
            arquivo_temp = baixar_planilha(arg)
            caminho = arquivo_temp
        else:
            # Argumento é um caminho local
            caminho = arg
    elif SHAREPOINT_LINK:
        # Usar link configurado no script
        arquivo_temp = baixar_planilha(SHAREPOINT_LINK)
        caminho = arquivo_temp
    else:
        # Fallback: usar arquivo local
        caminho = os.path.join(script_dir, "RelatorioManifestacoes_tratado.xlsx")
        print("[*] Nenhum link do SharePoint configurado, usando arquivo local")

    if not os.path.exists(caminho):
        print(f"Erro: arquivo não encontrado: {caminho}")
        sys.exit(1)

    print(f"[*] Lendo planilha: {caminho}")
    registros, headers = carregar_planilha(caminho)
    print(f"[*] {len(registros)} registros encontrados")

    print("[*] Processando dados...")
    dados = processar_dados(registros)

    # Salvar JSON na pasta public do React
    saida = os.path.join(script_dir, "public", "dados.json")
    with open(saida, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

    print(f"[OK] Arquivo gerado: {saida}")
    print(f"     Total: {dados['kpis']['total']} manifestações")
    print(f"     Concluídas: {dados['kpis']['concluidas']} ({dados['kpis']['percentual_concluidas']}%)")
    print(f"     Tempo médio de resposta: {dados['kpis']['media_dias_resposta']} dias")
    print(f"     Taxa no prazo: {dados['kpis']['taxa_no_prazo']}%")
    print(f"     Atualizado em: {dados['meta']['data_geracao']}")
    
    # Limpar arquivo temporário
    if arquivo_temp:
        try:
            os.unlink(arquivo_temp)
        except:
            pass


if __name__ == "__main__":
    main()
