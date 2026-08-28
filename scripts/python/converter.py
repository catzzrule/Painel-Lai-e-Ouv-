"""
Conversor de Planilha XLSX → JSON para o Painel da Ouvidoria
=============================================================
Uso: python converter.py
     python converter.py caminho/para/planilha.xlsx

Gera o arquivo 'dados.json' na pasta 'public'.
"""

import json
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Erro: módulo 'openpyxl' não encontrado.")
    print("Instale com: pip install openpyxl")
    sys.exit(1)


def carregar_planilha(caminho):
    """Carrega a planilha e retorna lista de dicionários."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    registros = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        registro = {}
        for i, valor in enumerate(row):
            if i < len(headers):
                chave = headers[i]
                if isinstance(valor, datetime):
                    registro[chave] = valor.strftime("%Y-%m-%d")
                else:
                    registro[chave] = valor
        registros.append(registro)

    return registros, headers


def processar_dados(registros):
    """Processa os registros e gera os dados agregados para o painel."""
    dados = {}

    # --- Metadados ---
    dados["meta"] = {
        "total_registros": len(registros),
        "data_geracao": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    # --- KPIs ---
    situacoes = Counter()
    decisoes = Counter()
    subtipos = Counter()
    status_prazo = Counter()
    recursos = Counter()
    generos = Counter()
    tipos_pessoa = Counter()
    estados = Counter()
    meses = Counter()
    
    # Para capturar as áreas mais demandadas
    areas_situacoes = defaultdict(Counter)
    areas_decisoes = defaultdict(Counter)
    areas_subtipos = defaultdict(Counter)
    areas_status = defaultdict(Counter)
    areas_recursos = defaultdict(Counter)

    dias_resposta_total = []
    dias_prazo_total = []
    dias_folga_total = []

    # Agregações mensais
    mensal_qtd = defaultdict(int)
    mensal_dias_resp = defaultdict(list)
    mensal_situacao = defaultdict(lambda: Counter())
    mensal_decisao = defaultdict(lambda: Counter())
    mensal_status_prazo = defaultdict(lambda: Counter())

    for r in registros:
        sit = r.get("Situação") or r.get("Situa\u00e7\u00e3o") or _get_key(r, "Situa")
        subtipo = r.get("Subtipo de Formulário") or r.get("Subtipo de Formul\u00e1rio") or _get_key(r, "Subtipo")
        decisao = r.get("Especificação Decisão") or r.get("Especifica\u00e7\u00e3o Decis\u00e3o") or _get_key(r, "Especifica")
        status = r.get("Status do Prazo") or _get_key(r, "Status do Prazo")
        recurso = r.get("Situação do Recurso") or r.get("Situa\u00e7\u00e3o do Recurso") or _get_key(r, "Recurso")
        genero = r.get("Gênero") or r.get("G\u00eanero") or _get_key(r, "nero")
        tipo_p = r.get("Tipo de Pessoa") or _get_key(r, "Tipo de Pessoa")
        estado = r.get("Estado") or _get_key(r, "Estado")
        ano_mes = r.get("Ano-Mês Cadastro") or r.get("Ano-M\u00eas Cadastro") or _get_key(r, "Ano-M")
        dias_resp = r.get("Dias para Resposta") or _get_key(r, "Dias para Resposta")
        dias_prazo = r.get("Dias de Prazo (SLA)") or _get_key(r, "Dias de Prazo")
        dias_folga = r.get("Dias de Folga/Atraso") or _get_key(r, "Dias de Folga")
        
        # Área Responsável
        resp = str(r.get("Responsável pela Resposta") or _get_key(r, "Respons") or "Não Identificada").strip()

        if sit:
            situacoes[str(sit)] += 1
            if resp and resp != "None": areas_situacoes[str(sit)][resp] += 1
        if subtipo:
            subtipos[str(subtipo)] += 1
            if resp and resp != "None": areas_subtipos[str(subtipo)][resp] += 1
        if decisao:
            decisoes[str(decisao)] += 1
            if resp and resp != "None": areas_decisoes[str(decisao)][resp] += 1
        if status:
            status_prazo[str(status)] += 1
            if resp and resp != "None": areas_status[str(status)][resp] += 1
        if recurso:
            recursos[str(recurso)] += 1
            if resp and resp != "None": areas_recursos[str(recurso)][resp] += 1
        if genero:
            generos[str(genero)] += 1
        if tipo_p:
            tipos_pessoa[str(tipo_p)] += 1
        if estado:
            estados[str(estado)] += 1
        if ano_mes:
            mes_str = str(ano_mes)
            meses[mes_str] += 1
            mensal_qtd[mes_str] += 1
            if sit:
                mensal_situacao[mes_str][str(sit)] += 1
            if decisao:
                mensal_decisao[mes_str][str(decisao)] += 1
            if status:
                mensal_status_prazo[mes_str][str(status)] += 1

        if dias_resp is not None:
            try:
                val = int(dias_resp)
                dias_resposta_total.append(val)
                if ano_mes:
                    mensal_dias_resp[str(ano_mes)].append(val)
            except (ValueError, TypeError):
                pass
        if dias_prazo is not None:
            try:
                dias_prazo_total.append(int(dias_prazo))
            except (ValueError, TypeError):
                pass
        if dias_folga is not None:
            try:
                dias_folga_total.append(int(dias_folga))
            except (ValueError, TypeError):
                pass

    # KPIs
    total = len(registros)

    concluidas = 0
    for k, v in situacoes.items():
        if "onclu" in k:
            concluidas += v

    em_aberto = 0
    for k in situacoes:
        if "Cadastrada" in k or "Encaminhada" in k:
            em_aberto += situacoes[k]

    com_recurso = 0
    for k in recursos:
        if "respondido" in k.lower():
            com_recurso += recursos[k]

    respondidas_prazo = 0
    for k in status_prazo:
        if "respondida" in k.lower() and "prazo" in k.lower():
            respondidas_prazo += status_prazo[k]

    no_prazo_total = 0
    for k in status_prazo:
        if "prazo" in k.lower():
            no_prazo_total += status_prazo[k]

    media_resp = round(sum(dias_resposta_total) / len(dias_resposta_total), 1) if dias_resposta_total else 0
    media_prazo = round(sum(dias_prazo_total) / len(dias_prazo_total), 1) if dias_prazo_total else 0
    media_folga = round(sum(dias_folga_total) / len(dias_folga_total), 1) if dias_folga_total else 0

    dados["kpis"] = {
        "total": total,
        "concluidas": concluidas,
        "percentual_concluidas": round((concluidas / total) * 100, 1) if total > 0 else 0,
        "em_aberto": em_aberto,
        "com_recurso": com_recurso,
        "respondidas_prazo": respondidas_prazo,
        "taxa_no_prazo": round((respondidas_prazo / concluidas) * 100, 1) if concluidas > 0 else 0,
        "no_prazo_total": no_prazo_total,
        "media_dias_resposta": media_resp,
        "media_dias_prazo_sla": media_prazo,
        "media_dias_folga": media_folga,
    }

    # --- Distribuições ---
    dados["situacoes"] = dict(situacoes.most_common())
    dados["decisoes"] = dict(decisoes.most_common())
    dados["subtipos"] = dict(subtipos.most_common())
    dados["status_prazo"] = dict(status_prazo.most_common())
    dados["recursos"] = dict(recursos.most_common())
    dados["generos"] = dict(generos.most_common())
    dados["tipos_pessoa"] = dict(tipos_pessoa.most_common())
    dados["estados"] = dict(sorted(estados.items(), key=lambda x: -x[1]))

    def extrair_top_area(counter_dict):
        resultado = {}
        for k, counter in counter_dict.items():
            top = counter.most_common(1)
            if top:
                resultado[k] = f"{top[0][0]} ({top[0][1]})"
            else:
                resultado[k] = "Não informada"
        return resultado

    dados["principais_areas"] = {
        "situacoes": extrair_top_area(areas_situacoes),
        "decisoes": extrair_top_area(areas_decisoes),
        "subtipos": extrair_top_area(areas_subtipos),
        "status_prazo": extrair_top_area(areas_status),
        "recursos": extrair_top_area(areas_recursos)
    }

    # --- Dados Mensais ---
    meses_ordenados = sorted(mensal_qtd.keys())
    dados["mensal"] = {
        "meses": meses_ordenados,
        "quantidades": [mensal_qtd[m] for m in meses_ordenados],
        "media_dias_resposta": [
            round(sum(mensal_dias_resp[m]) / len(mensal_dias_resp[m]), 1)
            if mensal_dias_resp[m]
            else 0
            for m in meses_ordenados
        ],
        "situacoes_por_mes": {m: dict(mensal_situacao[m]) for m in meses_ordenados},
        "status_prazo_por_mes": {m: dict(mensal_status_prazo[m]) for m in meses_ordenados},
    }

    # --- Lista detalhada (sem dados sensíveis) ---
    dados["registros"] = []
    for r in registros:
        registro_limpo = {}
        for chave, valor in r.items():
            # Omitir colunas sensíveis
            chave_lower = str(chave).lower() if chave else ""
            if any(
                termo in chave_lower
                for termo in ["usuário", "usuario", "responsável", "responsavel"]
            ):
                continue
            registro_limpo[chave] = valor
        dados["registros"].append(registro_limpo)

    return dados


def _get_key(registro, substring):
    """Busca uma chave no registro que contenha a substring (para lidar com encoding)."""
    for k, v in registro.items():
        if k and substring in str(k):
            return v
    return None


def main():
    # Determinar caminho da planilha
    if len(sys.argv) > 1:
        caminho = sys.argv[1]
    else:
        # Procurar na pasta do script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        caminho = os.path.join(script_dir, "RelatorioManifestacoes_tratado.xlsx")

    if not os.path.exists(caminho):
        print(f"Erro: arquivo não encontrado: {caminho}")
        sys.exit(1)

    print(f"[*] Lendo planilha: {caminho}")
    registros, headers = carregar_planilha(caminho)
    print(f"[*] {len(registros)} registros encontrados")

    print("[*] Processando dados...")
    dados = processar_dados(registros)

    # Salvar JSON na pasta public do React
    saida = os.path.join(os.path.dirname(os.path.abspath(__file__)), "public", "dados.json")
    with open(saida, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

    print(f"[OK] Arquivo gerado: {saida}")
    print(f"     Total: {dados['kpis']['total']} manifestacoes")
    print(f"     Concluidas: {dados['kpis']['concluidas']} ({dados['kpis']['percentual_concluidas']}%)")
    print(f"     Tempo medio de resposta: {dados['kpis']['media_dias_resposta']} dias")
    print(f"     Taxa no prazo: {dados['kpis']['taxa_no_prazo']}%")


if __name__ == "__main__":
    main()
