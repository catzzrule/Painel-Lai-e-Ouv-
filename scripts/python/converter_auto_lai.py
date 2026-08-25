"""
Conversor Automático: Painel LAI (Lei de Acesso à Informação)
=============================================================
Baixa a planilha LAI do SharePoint Online e gera 'public/dados_lai.json'.
Lê o link de compartilhamento a partir do arquivo .env (SHAREPOINT_URL_LAI).

Execução:
  python converter_auto_lai.py
"""

import json
import os
import sys
import tempfile
from collections import Counter, defaultdict
from datetime import datetime

# Carregar variáveis do .env com suporte a fallback nativo
def carregar_env():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        try:
            with open(env_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#") and "=" in line:
                        k, v = line.split("=", 1)
                        k = k.strip()
                        v = v.strip().strip('"').strip("'")
                        if k not in os.environ:
                            os.environ[k] = v
        except Exception as e:
            print(f"[!] Aviso ao ler .env: {e}")

carregar_env()

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    import requests
except ImportError:
    print("Erro: módulo 'requests' não encontrado. Instale com: pip install requests")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Erro: módulo 'openpyxl' não encontrado. Instale com: pip install openpyxl")
    sys.exit(1)


# Obter URL do .env ou usar fallback padrão
SHAREPOINT_LINK = os.environ.get(
    "SHAREPOINT_URL_LAI",
    "https://mdsgov-my.sharepoint.com/:x:/g/personal/rayssa_vitoria_esporte_gov_br/IQBvZ6Cbt41oRJBD1QuPbvLBAfoOSMBzDpwUbAT797m4ggA?e=FaCchc"
).strip()


def converter_link_sharepoint(link):
    if not link:
        return ""
    if "download.aspx" in link:
        return link
    if "/:x:/" in link or "/:w:/" in link or "/:p:/" in link or "sharepoint.com" in link:
        if "?" in link:
            return link + "&download=1"
        else:
            return link + "?download=1"
    return link


def baixar_planilha(link):
    print("[*] [LAI] Baixando planilha do SharePoint...")
    download_url = converter_link_sharepoint(link)
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }
    try:
        response = requests.get(download_url, headers=headers, allow_redirects=True, timeout=60)
        response.raise_for_status()

        content_type = response.headers.get("Content-Type", "")
        if "text/html" in content_type and len(response.content) < 5000:
            print("[!] O link retornou HTML em vez de um arquivo Excel.")
            sys.exit(1)

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.write(response.content)
        tmp.close()
        tamanho_kb = len(response.content) / 1024
        print(f"[OK] [LAI] Planilha baixada com sucesso ({tamanho_kb:.1f} KB)")
        return tmp.name
    except requests.exceptions.RequestException as e:
        print(f"[ERRO] [LAI] Falha ao baixar planilha: {e}")
        sys.exit(1)


def parse_date(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%Y %H:%M:%S", "%d/%m/%y", "%d/%m/%y %H:%M:%S", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


def carregar_planilha(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active

    # Detectar linha de cabeçalho
    header_row_idx = 1
    headers = []
    for r in range(1, 11):
        row_vals = [cell.value for cell in ws[r]]
        has_sit = any(isinstance(v, str) and ("Situa" in v) for v in row_vals)
        has_nup = any(isinstance(v, str) and ("Nup" in v) for v in row_vals)
        if has_sit and has_nup:
            header_row_idx = r
            headers = [str(v).strip() if v is not None else f"Coluna_{i}" for i, v in enumerate(row_vals)]
            break

    if not headers:
        header_row_idx = 1
        headers = [str(cell.value).strip() if cell.value is not None else f"Coluna_{i}" for i, cell in enumerate(ws[1])]

    def encontrar_coluna(termos):
        for h in headers:
            if any(t in h for t in termos):
                return h
        return None

    col_cadastro = encontrar_coluna(["Cadastro", "Data de Cadastro"])
    col_prazo = encontrar_coluna(["Prazo de Atendimento", "Prazo Recurso", "Prazo"])
    col_resposta = encontrar_coluna(["Data de Resposta", "Data Resposta"])
    col_situacao = encontrar_coluna(["Situação", "Situa"])
    col_nup = encontrar_coluna(["Nup", "NUP", "Protocolo"])

    tem_ano_mes = encontrar_coluna(["Ano-Mês Cadastro", "Ano-M"]) is not None
    tem_dias_resp = encontrar_coluna(["Dias para Resposta"]) is not None
    tem_dias_sla = encontrar_coluna(["Dias de Prazo (SLA)"]) is not None
    tem_dias_folga = encontrar_coluna(["Dias de Folga"]) is not None
    tem_status_prazo = encontrar_coluna(["Status do Prazo"]) is not None

    registros = []

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, len(headers) + 1)]
        if all(v is None for v in row_vals):
            continue

        registro = dict(zip(headers, row_vals))

        valid_ufs = {"AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO"}
        
        # Encontrar a chave do Estado
        chave_estado = None
        for k in headers:
            if k and "Estado" in str(k):
                chave_estado = k
                break
                
        estado_raw = str(registro.get(chave_estado) if chave_estado else "").strip().upper()
        if estado_raw not in valid_ufs:
            extracted_uf = "Não Informado"
            for val in row_vals[8:] if len(row_vals) > 8 else row_vals:
                if str(val).strip().upper() in valid_ufs:
                    extracted_uf = str(val).strip().upper()
                    break
            if chave_estado:
                registro[chave_estado] = extracted_uf
            else:
                registro["Estado"] = extracted_uf

        nup_val = registro.get(col_nup) if col_nup else None
        sit_val = registro.get(col_situacao) if col_situacao else None
        if nup_val is None or str(nup_val).strip() == "" or sit_val is None or str(sit_val).strip() == "":
            continue

        val_cadastro = parse_date(registro.get(col_cadastro)) if col_cadastro else None
        val_prazo = parse_date(registro.get(col_prazo)) if col_prazo else None
        val_resposta = parse_date(registro.get(col_resposta)) if col_resposta else None

        if col_cadastro and val_cadastro:
            registro[col_cadastro] = val_cadastro.strftime("%Y-%m-%d")
        if col_prazo and val_prazo:
            registro[col_prazo] = val_prazo.strftime("%Y-%m-%d")
        if col_resposta and val_resposta:
            registro[col_resposta] = val_resposta.strftime("%Y-%m-%d")

        if not tem_ano_mes and val_cadastro:
            registro["Ano-Mês Cadastro"] = val_cadastro.strftime("%Y-%m")

        if not tem_dias_resp:
            registro["Dias para Resposta"] = (val_resposta - val_cadastro).days if (val_resposta and val_cadastro) else None

        if not tem_dias_sla:
            registro["Dias de Prazo (SLA)"] = (val_prazo - val_cadastro).days if (val_prazo and val_cadastro) else None

        if not tem_dias_folga:
            registro["Dias de Folga/Atraso"] = (val_prazo - val_resposta).days if (val_prazo and val_resposta) else None

        if not tem_status_prazo:
            if val_resposta:
                registro["Status do Prazo"] = "Respondida no prazo" if (val_prazo and val_resposta <= val_prazo) else "Respondida fora do prazo"
            else:
                registro["Status do Prazo"] = "Em atraso" if (val_prazo and datetime.now() > val_prazo) else "Em aberto - no prazo"

        registros.append(registro)

    headers_finais = list(headers)
    for campo in ["Ano-Mês Cadastro", "Dias para Resposta", "Dias de Prazo (SLA)", "Dias de Folga/Atraso", "Status do Prazo"]:
        if campo not in headers_finais:
            headers_finais.append(campo)

    return registros, headers_finais


def _get_key(registro, substring):
    for k, v in registro.items():
        if k and substring in str(k):
            return v
    return None


def processar_dados(registros):
    dados = {}
    agora = datetime.now()
    dados["meta"] = {
        "painel": "LAI",
        "total_registros": len(registros),
        "data_geracao": agora.strftime("%Y-%m-%d %H:%M:%S"),
        "ultima_atualizacao": agora.strftime("%d/%m/%Y às %H:%M"),
    }

    situacoes = Counter()
    decisoes = Counter()
    subtipos = Counter()
    status_prazo = Counter()
    recursos = Counter()
    generos = Counter()
    tipos_pessoa = Counter()
    estados = Counter()
    meses = Counter()

    areas_situacoes = defaultdict(Counter)
    areas_decisoes = defaultdict(Counter)
    areas_subtipos = defaultdict(Counter)
    areas_status = defaultdict(Counter)
    areas_recursos = defaultdict(Counter)

    dias_resposta_total = []
    dias_prazo_total = []
    dias_folga_total = []

    mensal_qtd = defaultdict(int)
    mensal_dias_resp = defaultdict(list)
    mensal_situacao = defaultdict(lambda: Counter())
    mensal_decisao = defaultdict(lambda: Counter())
    mensal_status_prazo = defaultdict(lambda: Counter())

    for r in registros:
        sit = r.get("Situação") or r.get("Situa\u00e7\u00e3o") or _get_key(r, "Situa")
        subtipo = r.get("Subtipo de Formulário") or r.get("Subtipo de Formul\u00e1rio") or _get_key(r, "Subtipo")
        decisao = r.get("Especificação Decisão") or r.get("Especifica\u00e7\u00e3o Decis\u00e3o") or _get_key(r, "Especifica")
        status = r.get("Status do Prazo") or _get_key(r, "Status do Prazo")
        recurso = r.get("Situação do Recurso") or r.get("Situa\u00e7\u00e3o do Recurso") or _get_key(r, "Recurso")
        genero = r.get("Gênero") or r.get("G\u00eanero") or _get_key(r, "nero")
        tipo_p = r.get("Tipo de Pessoa") or _get_key(r, "Tipo de Pessoa")
        estado = r.get("Estado") or _get_key(r, "Estado")
        ano_mes = r.get("Ano-Mês Cadastro") or r.get("Ano-M\u00eas Cadastro") or _get_key(r, "Ano-M")
        dias_resp = r.get("Dias para Resposta") or _get_key(r, "Dias para Resposta")
        dias_prazo = r.get("Dias de Prazo (SLA)") or _get_key(r, "Dias de Prazo")
        dias_folga = r.get("Dias de Folga/Atraso") or _get_key(r, "Dias de Folga")
        resp = str(r.get("Responsável pela Resposta") or _get_key(r, "Respons") or "Não Identificada").strip()

        if sit:
            situacoes[str(sit)] += 1
            if resp and resp != "None": areas_situacoes[str(sit)][resp] += 1
        if subtipo:
            subtipos[str(subtipo)] += 1
            if resp and resp != "None": areas_subtipos[str(subtipo)][resp] += 1
        if decisao:
            decisoes[str(decisao)] += 1
            if resp and resp != "None": areas_decisoes[str(decisao)][resp] += 1
        if status:
            status_prazo[str(status)] += 1
            if resp and resp != "None": areas_status[str(status)][resp] += 1
        if recurso:
            recursos[str(recurso)] += 1
            if resp and resp != "None": areas_recursos[str(recurso)][resp] += 1
        if genero:
            generos[str(genero)] += 1
        if tipo_p:
            tipos_pessoa[str(tipo_p)] += 1
        if estado:
            estados[str(estado)] += 1
        if ano_mes:
            mes_str = str(ano_mes)
            meses[mes_str] += 1
            mensal_qtd[mes_str] += 1
            if sit: mensal_situacao[mes_str][str(sit)] += 1
            if decisao: mensal_decisao[mes_str][str(decisao)] += 1
            if status: mensal_status_prazo[mes_str][str(status)] += 1

        if dias_resp is not None:
            try:
                val = int(dias_resp)
                dias_resposta_total.append(val)
                if ano_mes: mensal_dias_resp[str(ano_mes)].append(val)
            except (ValueError, TypeError):
                pass
        if dias_prazo is not None:
            try: dias_prazo_total.append(int(dias_prazo))
            except (ValueError, TypeError): pass
        if dias_folga is not None:
            try: dias_folga_total.append(int(dias_folga))
            except (ValueError, TypeError): pass

    total = len(registros)
    concluidas = sum(v for k, v in situacoes.items() if "onclu" in k)
    em_aberto = sum(v for k, v in situacoes.items() if any(t in k for t in ["Cadastrada", "Prorrogada", "Encaminhada", "aberto"]))
    com_recurso = sum(v for k, v in recursos.items() if "respondido" in k.lower())
    respondidas_prazo = sum(v for k, v in status_prazo.items() if "respondida" in k.lower() and "prazo" in k.lower())
    no_prazo_total = sum(v for k, v in status_prazo.items() if "prazo" in k.lower())

    media_resp = round(sum(dias_resposta_total) / len(dias_resposta_total), 1) if dias_resposta_total else 0
    media_prazo = round(sum(dias_prazo_total) / len(dias_prazo_total), 1) if dias_prazo_total else 0
    media_folga = round(sum(dias_folga_total) / len(dias_folga_total), 1) if dias_folga_total else 0

    dados["kpis"] = {
        "total": total,
        "concluidas": concluidas,
        "percentual_concluidas": round((concluidas / total) * 100, 1) if total > 0 else 0,
        "em_aberto": em_aberto,
        "com_recurso": com_recurso,
        "respondidas_prazo": respondidas_prazo,
        "taxa_no_prazo": round((respondidas_prazo / concluidas) * 100, 1) if concluidas > 0 else 0,
        "no_prazo_total": no_prazo_total,
        "media_dias_resposta": media_resp,
        "media_dias_prazo_sla": media_prazo,
        "media_dias_folga": media_folga,
    }

    dados["situacoes"] = dict(situacoes.most_common())
    dados["decisoes"] = dict(decisoes.most_common())
    dados["subtipos"] = dict(subtipos.most_common())
    dados["status_prazo"] = dict(status_prazo.most_common())
    dados["recursos"] = dict(recursos.most_common())
    dados["generos"] = dict(generos.most_common())
    dados["tipos_pessoa"] = dict(tipos_pessoa.most_common())
    dados["estados"] = dict(sorted(estados.items(), key=lambda x: -x[1]))

    def extrair_top_area(counter_dict):
        resultado = {}
        for k, counter in counter_dict.items():
            top = counter.most_common(1)
            resultado[k] = f"{top[0][0]} ({top[0][1]})" if top else "Não informada"
        return resultado

    dados["principais_areas"] = {
        "situacoes": extrair_top_area(areas_situacoes),
        "decisoes": extrair_top_area(areas_decisoes),
        "subtipos": extrair_top_area(areas_subtipos),
        "status_prazo": extrair_top_area(areas_status),
        "recursos": extrair_top_area(areas_recursos)
    }

    meses_ordenados = sorted(mensal_qtd.keys())
    dados["mensal"] = {
        "meses": meses_ordenados,
        "quantidades": [mensal_qtd[m] for m in meses_ordenados],
        "media_dias_resposta": [
            round(sum(mensal_dias_resp[m]) / len(mensal_dias_resp[m]), 1) if mensal_dias_resp[m] else 0
            for m in meses_ordenados
        ],
        "situacoes_por_mes": {m: dict(mensal_situacao[m]) for m in meses_ordenados},
        "status_prazo_por_mes": {m: dict(mensal_status_prazo[m]) for m in meses_ordenados},
    }

    dados["registros"] = []
    for r in registros:
        registro_limpo = {}
        for chave, valor in r.items():
            chave_lower = str(chave).lower() if chave else ""
            if any(termo in chave_lower for termo in ["usuário", "usuario", "responsável", "responsavel"]):
                continue
            registro_limpo[chave] = valor
        dados["registros"].append(registro_limpo)

    return dados


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    arquivo_temp = None

    if len(sys.argv) > 1:
        arg = sys.argv[1]
        if arg.startswith("http"):
            arquivo_temp = baixar_planilha(arg)
            caminho = arquivo_temp
        else:
            caminho = arg
    elif SHAREPOINT_LINK:
        arquivo_temp = baixar_planilha(SHAREPOINT_LINK)
        caminho = arquivo_temp
    else:
        caminho = os.path.join(script_dir, "RelatorioManifestacoes_tratado.xlsx")
        print("[*] [LAI] Usando arquivo local")

    if not os.path.exists(caminho):
        print(f"[ERRO] [LAI] Arquivo não encontrado: {caminho}")
        sys.exit(1)

    print(f"[*] [LAI] Processando planilha: {caminho}")
    registros, _ = carregar_planilha(caminho)
    print(f"[*] [LAI] {len(registros)} registros encontrados")

    dados = processar_dados(registros)

    # Salvar em dados_lai.json e manter dados.json para retrocompatibilidade
    public_dir = os.path.join(script_dir, "public")
    os.makedirs(public_dir, exist_ok=True)

    saida_lai = os.path.join(public_dir, "dados_lai.json")
    with open(saida_lai, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

    saida_legado = os.path.join(public_dir, "dados.json")
    with open(saida_legado, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

    print(f"[OK] [LAI] Arquivo gerado com sucesso: {saida_lai}")
    print(f"     Total: {dados['kpis']['total']} pedidos")
    print(f"     Concluídos: {dados['kpis']['concluidas']} ({dados['kpis']['percentual_concluidas']}%)")
    print(f"     Taxa no prazo: {dados['kpis']['taxa_no_prazo']}%")

    if arquivo_temp:
        try:
            os.unlink(arquivo_temp)
        except Exception:
            pass


if __name__ == "__main__":
    main()
